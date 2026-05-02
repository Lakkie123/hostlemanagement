from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Count, Sum, Q
from django.utils import timezone
from datetime import date
# import random
#from django.core.mail import send_mail
#from django.conf import settings

from .models import (StudentProfile, Room, RoomType, RoomAllocation,
                     FeePayment, Complaint, Notice, Visitor, HostelFacility)
from .forms import (StudentRegistrationForm, LoginForm, RoomForm,
                    RoomAllocationForm, FeePaymentForm, ComplaintForm,
                    NoticeForm, VisitorForm, ProfileUpdateForm)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def is_admin(user):
    return user.is_staff or user.is_superuser


# ─── Public Views ──────────────────────────────────────────────────────────────
def home(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('admin_dashboard')
        return redirect('student_dashboard')
    
    facilities = HostelFacility.objects.filter(is_active=True)
    notices = Notice.objects.filter(is_active=True)[:5]
    total_rooms = Room.objects.count()
    available_rooms = Room.objects.filter(status='available').count()
    total_students = StudentProfile.objects.filter(status='active').count()
    return render(request, 'hostel_app/home.html', {
        'facilities': facilities,
        'notices': notices,
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'total_students': total_students,
    })

def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = StudentRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            login(request, user)  # ← yeh add karo
            messages.success(request, 'Registration successful! Please wait for admin approval.')
            return redirect('student_dashboard')  # ← yeh bhi change karo
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = StudentRegistrationForm()
    return render(request, 'hostel_app/register.html', {'form': form})

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if user.is_staff:
                return redirect('admin_dashboard')
            return redirect('student_dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = LoginForm()
    return render(request, 'hostel_app/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('home')


# ─── Student Views ─────────────────────────────────────────────────────────────

@login_required
def student_dashboard(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    try:
        profile = request.user.profile
    except StudentProfile.DoesNotExist:
        messages.error(request, 'Profile not found.')
        return redirect('home')

    allocation = RoomAllocation.objects.filter(student=profile, status='active').first()
    complaints = Complaint.objects.filter(student=profile).order_by('-created_at')[:5]
    notices = Notice.objects.filter(is_active=True).order_by('-created_at')[:5]
    payments = FeePayment.objects.filter(allocation__student=profile).order_by('-created_at')[:5] if allocation else []
    pending_fees = FeePayment.objects.filter(allocation__student=profile, status='pending').count() if allocation else 0

    return render(request, 'hostel_app/student/dashboard.html', {
        'profile': profile,
        'allocation': allocation,
        'complaints': complaints,
        'notices': notices,
        'payments': payments,
        'pending_fees': pending_fees,
    })


@login_required
def student_profile(request):
    try:
        profile = request.user.profile
    except StudentProfile.DoesNotExist:
        messages.error(request, 'Profile not found.')
        return redirect('home')
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=profile)
        if form.is_valid():
            profile.phone = form.cleaned_data['phone']
            profile.address = form.cleaned_data['address']
            profile.guardian_name = form.cleaned_data['guardian_name']
            profile.guardian_phone = form.cleaned_data['guardian_phone']
            profile.save()
            request.user.first_name = form.cleaned_data['first_name']
            request.user.last_name = form.cleaned_data['last_name']
            request.user.email = form.cleaned_data['email']
            request.user.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('student_profile')
    else:
        form = ProfileUpdateForm(instance=profile)
    return render(request, 'hostel_app/student/profile.html', {'profile': profile, 'form': form, 'profile_fields': []})


@login_required
def my_room(request):
    try:
        profile = request.user.profile
    except StudentProfile.DoesNotExist:
        return redirect('home')
    allocation = RoomAllocation.objects.filter(student=profile, status='active').first()
    return render(request, 'hostel_app/student/my_room.html', {'allocation': allocation, 'profile': profile})


@login_required
def submit_complaint(request):
    try:
        profile = request.user.profile
    except StudentProfile.DoesNotExist:
        return redirect('home')
    if request.method == 'POST':
        form = ComplaintForm(request.POST)
        if form.is_valid():
            complaint = form.save(commit=False)
            complaint.student = profile
            complaint.save()
            messages.success(request, 'Complaint submitted successfully!')
            return redirect('my_complaints')
    else:
        form = ComplaintForm()
    return render(request, 'hostel_app/student/complaint_form.html', {'form': form})


@login_required
def my_complaints(request):
    try:
        profile = request.user.profile
    except StudentProfile.DoesNotExist:
        return redirect('home')
    complaints = Complaint.objects.filter(student=profile).order_by('-created_at')
    return render(request, 'hostel_app/student/my_complaints.html', {'complaints': complaints})


@login_required
def my_payments(request):
    try:
        profile = request.user.profile
    except StudentProfile.DoesNotExist:
        return redirect('home')
    payments = FeePayment.objects.filter(allocation__student=profile).order_by('-created_at')
    return render(request, 'hostel_app/student/my_payments.html', {'payments': payments, 'profile': profile})

@login_required
def pay_fee(request):
    try:
        profile = request.user.profile
    except:
        return redirect('home')
    import calendar
    from datetime import date
    allocation = RoomAllocation.objects.filter(student=profile, status='active').first()
    pending_payments = FeePayment.objects.filter(allocation__student=profile, status='pending').order_by('-created_at')
    recent_payments = FeePayment.objects.filter(allocation__student=profile).order_by('-created_at')[:6]
    months = []
    today = date.today()
    for i in range(-1, 6):
        m = (today.month - 1 + i) % 12 + 1
        y = today.year + ((today.month - 1 + i) // 12)
        months.append(f"{calendar.month_name[m]} {y}")
    return render(request, 'hostel_app/student/pay_fee.html', {
        'profile': profile, 'allocation': allocation,
        'pending_payments': pending_payments,
        'recent_payments': recent_payments, 'months': months,
    })

@login_required
def confirm_payment(request):
    if request.method != 'POST':
        return redirect('pay_fee')
    try:
        profile = request.user.profile
    except:
        return redirect('home')
    from datetime import date
    allocation = get_object_or_404(RoomAllocation, pk=request.POST.get('allocation_id'), student=profile)
    
    payment = FeePayment(
        allocation=allocation,
        amount=request.POST.get('amount', allocation.monthly_rent),
        payment_date=date.today(),
        due_date=date.today(),
        month=request.POST.get('month', ''),
        status='pending',
        payment_method=request.POST.get('payment_method', 'upi'),
        transaction_id=request.POST.get('transaction_id', ''),
        notes=request.POST.get('notes', ''),
        received_by=None,
    )
    # Slip image save
    if 'slip_image' in request.FILES:
        payment.slip_image = request.FILES['slip_image']
    payment.save()
    
    messages.success(request, '✅ Payment submit ho gayi! Admin slip verify karke confirm karega.')
    return redirect('pay_fee')


@login_required
def add_visitor(request):
    try:
        profile = request.user.profile
    except StudentProfile.DoesNotExist:
        return redirect('home')
    if request.method == 'POST':
        form = VisitorForm(request.POST)
        if form.is_valid():
            visitor = form.save(commit=False)
            visitor.student = profile
            visitor.save()
            messages.success(request, 'Visitor registered successfully!')
            return redirect('my_visitors')
    else:
        form = VisitorForm()
    return render(request, 'hostel_app/student/visitor_form.html', {'form': form})


@login_required
def my_visitors(request):
    try:
        profile = request.user.profile
    except StudentProfile.DoesNotExist:
        return redirect('home')
    visitors = Visitor.objects.filter(student=profile).order_by('-check_in')
    return render(request, 'hostel_app/student/my_visitors.html', {'visitors': visitors})


# ─── Admin Views ───────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    total_students = StudentProfile.objects.filter(status='active').count()
    total_rooms = Room.objects.count()
    available_rooms = Room.objects.filter(status='available').count()
    occupied_rooms = Room.objects.filter(status='occupied').count()
    pending_students = StudentProfile.objects.filter(status='pending').count()
    open_complaints = Complaint.objects.filter(status='open').count()
    total_revenue = FeePayment.objects.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or 0
    pending_fees = FeePayment.objects.filter(status='pending').count()
    recent_allocations = RoomAllocation.objects.filter(status='active').order_by('-created_at')[:5]
    recent_complaints = Complaint.objects.order_by('-created_at')[:5]
    recent_payments = FeePayment.objects.order_by('-created_at')[:5]

    room_status_data = {
        'available': available_rooms,
        'occupied': occupied_rooms,
        'maintenance': Room.objects.filter(status='maintenance').count(),
        'reserved': Room.objects.filter(status='reserved').count(),
    }

    return render(request, 'hostel_app/admin/dashboard.html', {
        'total_students': total_students,
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'occupied_rooms': occupied_rooms,
        'pending_students': pending_students,
        'open_complaints': open_complaints,
        'total_revenue': total_revenue,
        'pending_fees': pending_fees,
        'recent_allocations': recent_allocations,
        'recent_complaints': recent_complaints,
        'recent_payments': recent_payments,
        'room_status_data': room_status_data,
    })


@login_required
@user_passes_test(is_admin)
def manage_rooms(request):
    rooms = Room.objects.select_related('room_type').all().order_by('room_number')
    status_filter = request.GET.get('status', '')
    floor_filter = request.GET.get('floor', '')
    if status_filter:
        rooms = rooms.filter(status=status_filter)
    if floor_filter:
        rooms = rooms.filter(floor=floor_filter)
    return render(request, 'hostel_app/admin/rooms.html', {
        'rooms': rooms,
        'status_filter': status_filter,
        'floor_filter': floor_filter,
        'status_choices': Room.STATUS_CHOICES,
        'floor_choices': Room.FLOOR_CHOICES,
    })


@login_required
@user_passes_test(is_admin)
def add_room(request):
    if request.method == 'POST':
        form = RoomForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Room added successfully!')
            return redirect('manage_rooms')
    else:
        form = RoomForm()
    return render(request, 'hostel_app/admin/room_form.html', {'form': form, 'title': 'Add Room'})


@login_required
@user_passes_test(is_admin)
def edit_room(request, pk):
    room = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        form = RoomForm(request.POST, instance=room)
        if form.is_valid():
            form.save()
            messages.success(request, f'Room {room.room_number} updated successfully!')
            return redirect('manage_rooms')
    else:
        form = RoomForm(instance=room)
    return render(request, 'hostel_app/admin/room_form.html', {'form': form, 'title': f'Edit Room {room.room_number}', 'room': room})


@login_required
@user_passes_test(is_admin)
def room_detail(request, pk):
    room = get_object_or_404(Room, pk=pk)
    allocations = RoomAllocation.objects.filter(room=room).select_related('student__user').order_by('-created_at')
    return render(request, 'hostel_app/admin/room_detail.html', {'room': room, 'allocations': allocations})


@login_required
@user_passes_test(is_admin)
def available_rooms(request):
    rooms = Room.objects.filter(status='available').select_related('room_type')
    return render(request, 'hostel_app/admin/available_rooms.html', {'rooms': rooms})

@login_required
@user_passes_test(is_admin)
def manage_students(request):
    students = StudentProfile.objects.select_related('user').all().order_by('-created_at')
    status_filter = request.GET.get('status', '')
    gender_filter = request.GET.get('gender', '')
    search = request.GET.get('search', '')
    if status_filter:
        students = students.filter(status=status_filter)
    if gender_filter:
        students = students.filter(gender=gender_filter)
    if search:
        students = students.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(roll_number__icontains=search) |
            Q(user__email__icontains=search)
        )
    return render(request, 'hostel_app/admin/students.html', {
        'students': students,
        'status_filter': status_filter,
        'gender_filter': gender_filter,
        'search': search,
        'boys_count': StudentProfile.objects.filter(gender='M').count(),
        'girls_count': StudentProfile.objects.filter(gender='F').count(),
    })
    
    
@login_required
@user_passes_test(is_admin)
def student_detail(request, pk):
    student = get_object_or_404(StudentProfile, pk=pk)
    allocations = RoomAllocation.objects.filter(student=student).order_by('-created_at')
    complaints = Complaint.objects.filter(student=student).order_by('-created_at')
    payments = FeePayment.objects.filter(allocation__student=student).order_by('-created_at')
    return render(request, 'hostel_app/admin/student_detail.html', {
        'student': student,
        'allocations': allocations,
        'complaints': complaints,
        'payments': payments,
        
    })

@login_required
@user_passes_test(is_admin)
def approve_student(request, pk):
    student = get_object_or_404(StudentProfile, pk=pk)
    student.status = 'active'
    student.save()
    messages.success(request, f'{student.user.get_full_name()} approved successfully!')
    return redirect('manage_students')

@login_required
@user_passes_test(is_admin)
def reject_student(request, pk):
    student = get_object_or_404(StudentProfile, pk=pk)
    user = student.user
    name = user.get_full_name()
    student.delete()
    user.delete()
    messages.success(request, f'{name} rejected and removed successfully!')
    return redirect('manage_students')
 
@login_required
@user_passes_test(is_admin)
def bulk_delete_students(request):
    if request.method == 'POST':
        selected = request.POST.getlist('selected_students')
        if selected:
            students = StudentProfile.objects.filter(pk__in=selected)
            users = [s.user for s in students]
            students.delete()
            for user in users:
                user.delete()
            messages.success(request, f'{len(selected)} students deleted successfully!')
        else:
            messages.error(request, 'Koi student select nahi kiya!')
    return redirect('manage_students') 


@login_required
@user_passes_test(is_admin)
def manage_allocations(request):
    allocations = RoomAllocation.objects.select_related('student__user', 'room').order_by('-created_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        allocations = allocations.filter(status=status_filter)
    return render(request, 'hostel_app/admin/allocations.html', {
        'allocations': allocations,
        'status_filter': status_filter,
    })


@login_required
@user_passes_test(is_admin)
def add_allocation(request):
    if request.method == 'POST':
        form = RoomAllocationForm(request.POST)
        if form.is_valid():
            allocation = form.save(commit=False)
            allocation.allocated_by = request.user
            allocation.save()
            # Update room occupancy
            room = allocation.room
            room.current_occupancy += 1
            if room.current_occupancy >= room.room_type.capacity:
                room.status = 'occupied'
            room.save()
            messages.success(request, 'Room allocated successfully!')
            return redirect('manage_allocations')
    else:
        form = RoomAllocationForm()
    return render(request, 'hostel_app/admin/allocation_form.html', {'form': form, 'title': 'New Allocation'})


@login_required
@user_passes_test(is_admin)
def vacate_allocation(request, pk):
    allocation = get_object_or_404(RoomAllocation, pk=pk)
    allocation.status = 'vacated'
    allocation.actual_vacate_date = date.today()
    allocation.save()
    room = allocation.room
    room.current_occupancy = max(0, room.current_occupancy - 1)
    if room.current_occupancy < room.room_type.capacity:
        room.status = 'available'
    room.save()
    messages.success(request, 'Room vacated successfully!')
    return redirect('manage_allocations')


@login_required
@user_passes_test(is_admin)
def manage_fees(request):
    payments = FeePayment.objects.select_related('allocation__student__user').order_by('-created_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        payments = payments.filter(status=status_filter)
    total_collected = FeePayment.objects.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or 0
    total_pending = FeePayment.objects.filter(status='pending').aggregate(Sum('amount'))['amount__sum'] or 0
    return render(request, 'hostel_app/admin/fees.html', {
        'payments': payments,
        'status_filter': status_filter,
        'total_collected': total_collected,
        'total_pending': total_pending,
    })


@login_required
@user_passes_test(is_admin)
def add_fee(request):
    if request.method == 'POST':
        form = FeePaymentForm(request.POST)
        if form.is_valid():
            fee = form.save(commit=False)
            fee.received_by = request.user
            fee.save()
            messages.success(request, 'Fee record added successfully!')
            return redirect('manage_fees')
    else:
        form = FeePaymentForm()
    return render(request, 'hostel_app/admin/fee_form.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def manage_complaints(request):
    complaints = Complaint.objects.select_related('student__user').order_by('-created_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    return render(request, 'hostel_app/admin/complaints.html', {
        'complaints': complaints,
        'status_filter': status_filter,
    })


@login_required
@user_passes_test(is_admin)
def resolve_complaint(request, pk):
    complaint = get_object_or_404(Complaint, pk=pk)
    if request.method == 'POST':
        complaint.status = request.POST.get('status', 'resolved')
        complaint.resolution_notes = request.POST.get('resolution_notes', '')
        complaint.resolved_by = request.user
        complaint.resolved_at = timezone.now()
        complaint.save()
        messages.success(request, 'Complaint updated!')
        return redirect('manage_complaints')
    return render(request, 'hostel_app/admin/resolve_complaint.html', {'complaint': complaint})


@login_required
@user_passes_test(is_admin)
def manage_notices(request):
    notices = Notice.objects.all().order_by('-created_at')
    return render(request, 'hostel_app/admin/notices.html', {'notices': notices})


@login_required
@user_passes_test(is_admin)
def add_notice(request):
    if request.method == 'POST':
        form = NoticeForm(request.POST)
        if form.is_valid():
            notice = form.save(commit=False)
            notice.posted_by = request.user
            notice.save()
            messages.success(request, 'Notice posted successfully!')
            return redirect('manage_notices')
    else:
        form = NoticeForm()
    return render(request, 'hostel_app/admin/notice_form.html', {'form': form, 'title': 'Post Notice'})


@login_required
@user_passes_test(is_admin)
def edit_notice(request, pk):
    notice = get_object_or_404(Notice, pk=pk)
    if request.method == 'POST':
        form = NoticeForm(request.POST, instance=notice)
        if form.is_valid():
            form.save()
            messages.success(request, 'Notice updated!')
            return redirect('manage_notices')
    else:
        form = NoticeForm(instance=notice)
    return render(request, 'hostel_app/admin/notice_form.html', {'form': form, 'title': 'Edit Notice', 'notice': notice})


@login_required
@user_passes_test(is_admin)
def manage_visitors(request):
    visitors = Visitor.objects.select_related('student__user').order_by('-check_in')
    return render(request, 'hostel_app/admin/visitors.html', {'visitors': visitors})


# ─── API Endpoints ─────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_admin)
def room_stats_api(request):
    data = {
        'available': Room.objects.filter(status='available').count(),
        'occupied': Room.objects.filter(status='occupied').count(),
        'maintenance': Room.objects.filter(status='maintenance').count(),
        'reserved': Room.objects.filter(status='reserved').count(),
    }
    return JsonResponse(data)

@login_required
@user_passes_test(is_admin)
def verify_payment(request, pk):
    payment = get_object_or_404(FeePayment, pk=pk)
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'approve':
            payment.status = 'paid'
            payment.received_by = request.user
            payment.save()
            messages.success(request, f'✅ ₹{payment.amount} verify ho gayi — {payment.allocation.student.user.get_full_name()}')
        elif action == 'reject':
            reject_reason = request.POST.get('reject_reason', '')
            payment.status = 'overdue'
            payment.notes = f'REJECTED: {reject_reason}'
            payment.transaction_id = ''
            payment.save()
            messages.error(request, f'❌ Payment reject kar di — {payment.allocation.student.user.get_full_name()}')
    return redirect('pending_verifications')


@login_required  
@user_passes_test(is_admin)
def pending_verifications(request):
    payments = FeePayment.objects.filter(
        status='pending', transaction_id__gt=''
    ).select_related('allocation__student__user', 'allocation__room').order_by('-created_at')
    return render(request, 'hostel_app/admin/pending_verifications.html', {'payments': payments})


def about(request):
    previews = [
        ('🛏️', 'Hostel Rooms'),
        ('📚', 'Study Area'),
        ('🏏', 'Sports Ground'),
        ('🍽️', 'Mess Hall'),
        ('🎭', 'Cultural Events'),
        ('🏛️', 'Campus Building'),
    ]
    return render(request, 'hostel_app/about.html', {'previews': previews})

def gallery(request):
    return render(request, 'hostel_app/gallery.html')




@login_required
@user_passes_test(is_admin)
def export_students(request):
    wb = openpyxl.Workbook()

    NAVY = "0A1628"
    GOLD = "C9A227"
    PINK = "C2185B"
    BLUE = "1A5FA8"
    WHITE = "FFFFFF"
    LIGHT_BLUE = "E3F2FD"
    LIGHT_PINK = "FCE4EC"
    GRAY = "F4F6FA"

    def make_border():
        thin = Side(style='thin', color="E8ECF4")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell, bg_color):
        cell.font = Font(bold=True, color=WHITE, name='Arial', size=11)
        cell.fill = PatternFill("solid", start_color=bg_color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = make_border()

    def style_data(cell, bg_color=None):
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = make_border()
        if bg_color:
            cell.fill = PatternFill("solid", start_color=bg_color)

    def write_sheet(ws, students, title, header_color, row_color):
        ws.row_dimensions[1].height = 50
        ws.row_dimensions[2].height = 35

        ws.merge_cells('A1:K1')
        tc = ws['A1']
        tc.value = f"Government Polytechnic Dehradun — {title}"
        tc.font = Font(bold=True, color=WHITE, name='Arial', size=14)
        tc.fill = PatternFill("solid", start_color=NAVY)
        tc.alignment = Alignment(horizontal='center', vertical='center')

        headers = ['S.No', 'Full Name', 'Roll Number', 'Course', 'Year',
                   'Gender', 'Phone', 'Guardian Name', 'Guardian Phone', 'Status', 'Joined']
        widths = [6, 22, 15, 20, 6, 10, 14, 20, 14, 10, 14]

        for col, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=2, column=col, value=h)
            style_header(cell, header_color)
            ws.column_dimensions[get_column_letter(col)].width = w

        for i, student in enumerate(students, 1):
            row = i + 2
            ws.row_dimensions[row].height = 22
            bg = row_color if i % 2 == 0 else None
            data = [
                i,
                student.user.get_full_name(),
                student.roll_number,
                student.course,
                student.year_of_study,
                student.get_gender_display(),
                student.phone,
                student.guardian_name,
                student.guardian_phone,
                student.get_status_display(),
                student.created_at.strftime('%d %b %Y') if student.created_at else '',
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                style_data(cell, bg)
                if col == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        total_row = len(students) + 3
        ws.merge_cells(f'A{total_row}:K{total_row}')
        sc = ws[f'A{total_row}']
        sc.value = f'Total: {len(students)} Students'
        sc.font = Font(bold=True, color=NAVY, name='Arial', size=11)
        sc.fill = PatternFill("solid", start_color=GOLD)
        sc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[total_row].height = 25

    boys = list(StudentProfile.objects.filter(gender='M').select_related('user').order_by('roll_number'))
    girls = list(StudentProfile.objects.filter(gender='F').select_related('user').order_by('roll_number'))
    all_students = list(StudentProfile.objects.select_related('user').order_by('gender', 'roll_number'))

    ws_all = wb.active
    ws_all.title = "All Students"
    write_sheet(ws_all, all_students, "All Students", NAVY, GRAY)

    ws_boys = wb.create_sheet("Boys Hostel")
    write_sheet(ws_boys, boys, "Boys Hostel", BLUE, LIGHT_BLUE)

    ws_girls = wb.create_sheet("Girls Hostel")
    write_sheet(ws_girls, girls, "Girls Hostel", PINK, LIGHT_PINK)

    from datetime import date
    filename = f"Students_{date.today().strftime('%d_%b_%Y')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response